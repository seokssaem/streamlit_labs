# ============================================================
#  📰 네이버 IT 뉴스 크롤링 + 엑셀 저장
#  오늘 수업 목표: openpyxl로 크롤링 결과를 엑셀에 저장하기
# ============================================================

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ============================================================
#  STEP 1. 지난 시간 복습 — 뉴스 크롤링
# ============================================================

url = 'https://news.naver.com/section/105'
response = requests.get(url, headers={
    'User-Agent': 'Mozilla/5.0'
})
html = response.text
soup = BeautifulSoup(html, 'html.parser')

news_list = soup.find_all('strong', {'class': 'sa_text_strong'})

print(f'수집된 기사 수: {len(news_list)}개')
print('=' * 50)

# 지난 시간: 그냥 출력만 했음
for i, tag in enumerate(news_list, start=1):
    print(i, tag.text.strip())

print('=' * 50)


# ============================================================
#  STEP 2. 오늘 새로운 것 — 딕셔너리 리스트로 모으기
#  (엑셀에 저장하려면 데이터를 먼저 잘 정리해야 해요!)
# ============================================================

results = []    # 빈 리스트 준비

for i, tag in enumerate(news_list, start=1):
    제목 = tag.text.strip()

    # 부모 <a> 태그에서 링크 꺼내기
    parent_a = tag.find_parent('a')
    링크 = parent_a['href'] if parent_a else ''

    # 카드 영역에서 언론사 꺼내기
    card = tag.find_parent('li') or tag.find_parent('div')
    press_tag = card.find('div', {'class': 'sa_text_press'}) if card else None
    언론사 = press_tag.text.strip() if press_tag else ''

    results.append({        # 딕셔너리를 리스트에 하나씩 추가
        'N'   : i,
        '제목' : 제목,
        '언론사': 언론사,
        '링크' : 링크,
    })

print('results 미리보기:')
for r in results[:3]:       # 앞 3개만 확인
    print(r)
print('=' * 50)


# ============================================================
#  STEP 3. openpyxl 기초 — 워크북/시트 만들기
# ============================================================

# ① 워크북(=엑셀 파일) 만들기
wb = Workbook()

# ② 활성 시트 가져오기
ws = wb.active

# ③ 시트 이름 바꾸기
ws.title = '네이버 IT뉴스'

print('워크북 생성:', wb)
print('시트 이름:', ws.title)
print('=' * 50)


# ============================================================
#  STEP 4. openpyxl — 셀에 데이터 입력하기
# ============================================================

# 방법 1: 셀 주소로 직접 입력
ws['A1'] = 'N'
ws['B1'] = '제목'
ws['C1'] = '언론사'
ws['D1'] = '링크'

# 방법 2: append() — 한 행씩 통째로 추가 (더 편리!)
# ws.append(['N', '제목', '언론사', '링크'])    # ← 이렇게도 됩니다


# ============================================================
#  STEP 5. openpyxl 스타일 — 제목 행 꾸미기
# ============================================================

# 파란 배경
header_fill = PatternFill('solid', fgColor='1F6FEB')

# 흰색 굵은 글씨
header_font = Font(bold=True, color='FFFFFF')

# 가운데 정렬
center = Alignment(horizontal='center')

# A1~D1 에 스타일 적용
for cell in ws[1]:      # ws[1] = 1번 행 전체
    cell.fill      = header_fill
    cell.font      = header_font
    cell.alignment = center


# ============================================================
#  STEP 6. openpyxl — 크롤링 결과를 엑셀에 추가하기
# ============================================================

for row in results:
    ws.append([row['N'], row['제목'], row['언론사'], row['링크']])
    #          ↑ append()에 리스트로 넘기면 한 행씩 추가됩니다!

print(f'엑셀에 입력된 행 수: {ws.max_row}행')   # 제목행 포함
print('=' * 50)


# ============================================================
#  STEP 7. openpyxl — 열 너비 조정 & 저장
# ============================================================

# 열 너비 조정
ws.column_dimensions['A'].width = 5    # N
ws.column_dimensions['B'].width = 60   # 제목 (넓게!)
ws.column_dimensions['C'].width = 15   # 언론사
ws.column_dimensions['D'].width = 50   # 링크

# 엑셀 파일로 저장!
wb.save('네이버_IT뉴스.xlsx')

print('✅ 엑셀 저장 완료! → 네이버_IT뉴스.xlsx')
print(f'   총 {len(results)}개 기사가 저장되었습니다.')
