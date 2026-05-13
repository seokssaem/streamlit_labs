"""
============================================================
  📰 네이버 IT 뉴스 크롤러 — GUI 버전 (보너스 시연용)
  "openpyxl + 크롤링을 GUI로 만들면 이렇게 됩니다!"
============================================================
  설치:
    pip install PyQt6 requests beautifulsoup4 openpyxl
============================================================
"""

import sys
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget,
    QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel,
    QTableWidget, QTableWidgetItem,
    QHeaderView, QStatusBar, QMessageBox,
    QFileDialog, QProgressBar,
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QColor, QPalette


# ── 크롤링 함수 (수업 때 배운 코드 그대로!) ─────────────────

def crawl_naver_it_news():
    url = 'https://news.naver.com/section/105'
    response = requests.get(url, headers={'User-Agent': 'Mozilla/5.0'})
    soup = BeautifulSoup(response.text, 'html.parser')

    news_list = soup.find_all('strong', {'class': 'sa_text_strong'})

    results = []
    for i, tag in enumerate(news_list, start=1):
        제목 = tag.text.strip()
        parent_a = tag.find_parent('a')
        링크 = parent_a['href'] if parent_a else ''
        card = tag.find_parent('li') or tag.find_parent('div')
        press_tag = card.find('a', {'class': 'sa_text_press'}) if card else None
        언론사 = press_tag.text.strip() if press_tag else ''

        results.append({'N': i, '제목': 제목, '언론사': 언론사, '링크': 링크})

    return results


# ── 엑셀 저장 함수 (수업 때 배운 코드 그대로!) ──────────────

def save_to_excel(results, filepath):
    wb = Workbook()
    ws = wb.active
    ws.title = '네이버 IT뉴스'

    ws.append(['N', '제목', '언론사', '링크'])

    for cell in ws[1]:
        cell.fill      = PatternFill('solid', fgColor='1F6FEB')
        cell.font      = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')

    for row in results:
        ws.append([row['N'], row['제목'], row['언론사'], row['링크']])

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 50

    wb.save(filepath)


# ── 백그라운드 크롤링 (화면 안 멈추게) ──────────────────────

class CrawlWorker(QThread):
    finished = pyqtSignal(list)
    error    = pyqtSignal(str)

    def run(self):
        try:
            self.finished.emit(crawl_naver_it_news())
        except Exception as e:
            self.error.emit(str(e))


# ── 스타일 ───────────────────────────────────────────────────

STYLE = """
QMainWindow, QWidget {
    background-color: #0d1117;
    color: #e6edf3;
    font-family: "Malgun Gothic", "AppleGothic", sans-serif;
    font-size: 13px;
}
QPushButton {
    background:#21262d; color:#e6edf3;
    border:1px solid #30363d; border-radius:8px;
    padding:10px 22px; font-size:14px; font-weight:bold;
}
QPushButton:hover   { background:#30363d; border-color:#58a6ff; color:#58a6ff; }
QPushButton:pressed { background:#1f6feb; color:white; border-color:#1f6feb; }
QPushButton:disabled{ background:#161b22; color:#484f58; }
QPushButton#btn_crawl {
    background:#1f6feb; border:none; color:white;
    font-size:15px; padding:12px 36px; border-radius:10px;
}
QPushButton#btn_crawl:hover    { background:#388bfd; }
QPushButton#btn_crawl:disabled { background:#0d419d; color:#8b949e; }
QPushButton#btn_save {
    background:#238636; border:none; color:white;
    border-radius:8px; padding:10px 24px;
}
QPushButton#btn_save:hover    { background:#2ea043; }
QPushButton#btn_save:disabled { background:#0f2d1a; color:#8b949e; }
QTableWidget {
    background:#161b22; border:1px solid #30363d;
    border-radius:8px; gridline-color:#21262d; color:#e6edf3;
}
QTableWidget::item:selected  { background:#1f6feb; color:white; }
QTableWidget::item:alternate { background:#1c2128; }
QHeaderView::section {
    background:#21262d; color:#58a6ff;
    padding:8px; border:none; font-weight:bold;
}
QProgressBar {
    background:#21262d; border:none; border-radius:4px; height:6px;
}
QProgressBar::chunk { background:#1f6feb; border-radius:4px; }
QStatusBar { background:#161b22; color:#8b949e; border-top:1px solid #30363d; }
QScrollBar:vertical { background:#0d1117; width:8px; border-radius:4px; }
QScrollBar::handle:vertical { background:#30363d; border-radius:4px; min-height:24px; }
QScrollBar::handle:vertical:hover { background:#58a6ff; }
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height:0; }
"""


# ── 메인 윈도우 ──────────────────────────────────────────────

class NewsApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.results = []
        self.setWindowTitle('📰 네이버 IT 뉴스 크롤러')
        self.setMinimumSize(980, 620)
        self.setStyleSheet(STYLE)
        self._init_ui()
        self.status = QStatusBar()
        self.setStatusBar(self.status)
        self.status.showMessage('👋  [크롤링 시작] 버튼을 눌러보세요!')

    def _init_ui(self):
        center = QWidget()
        self.setCentralWidget(center)
        lay = QVBoxLayout(center)
        lay.setSpacing(12)
        lay.setContentsMargins(20, 16, 20, 16)

        # 헤더
        title = QLabel('📰  네이버 IT 뉴스 크롤러')
        title.setStyleSheet('font-size:20px; font-weight:bold; color:#58a6ff;')
        sub = QLabel('수업 시간에 배운 크롤링 + openpyxl 을 GUI로 연결하면?')
        sub.setStyleSheet('font-size:12px; color:#8b949e;')
        lay.addWidget(title)
        lay.addWidget(sub)

        # 버튼 바
        btn_row = QHBoxLayout()
        self.btn_crawl = QPushButton('🕷️  크롤링 시작')
        self.btn_crawl.setObjectName('btn_crawl')
        self.btn_crawl.clicked.connect(self._start_crawl)

        self.btn_save = QPushButton('💾  엑셀로 저장')
        self.btn_save.setObjectName('btn_save')
        self.btn_save.setEnabled(False)
        self.btn_save.clicked.connect(self._save_excel)

        self.lbl_count = QLabel('기사 수: —')
        self.lbl_count.setStyleSheet(
            'font-size:14px; font-weight:bold; color:#3fb950;')

        btn_row.addWidget(self.btn_crawl)
        btn_row.addWidget(self.btn_save)
        btn_row.addStretch()
        btn_row.addWidget(self.lbl_count)
        lay.addLayout(btn_row)

        # 진행바
        self.pbar = QProgressBar()
        self.pbar.setVisible(False)
        self.pbar.setFixedHeight(6)
        self.pbar.setRange(0, 0)
        lay.addWidget(self.pbar)

        # 결과 테이블
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(['N', '제목', '언론사', '링크'])
        self.table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(
            QTableWidget.SelectionBehavior.SelectRows)
        self.table.verticalHeader().setVisible(False)
        self.table.verticalHeader().setDefaultSectionSize(30)
        lay.addWidget(self.table, 1)

    def _start_crawl(self):
        self.btn_crawl.setEnabled(False)
        self.btn_save.setEnabled(False)
        self.pbar.setVisible(True)
        self.table.setRowCount(0)
        self.lbl_count.setText('크롤링 중...')
        self.status.showMessage('🔄  네이버 IT 뉴스 가져오는 중...')

        self.worker = CrawlWorker()
        self.worker.finished.connect(self._on_done)
        self.worker.error.connect(self._on_error)
        self.worker.start()

    def _on_done(self, results):
        self.results = results
        self.pbar.setVisible(False)
        self.btn_crawl.setEnabled(True)
        self.btn_save.setEnabled(True)

        self.table.setRowCount(len(results))
        for r, row in enumerate(results):
            for c, key in enumerate(['N', '제목', '언론사', '링크']):
                item = QTableWidgetItem(str(row[key]))
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                if key == 'N':
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.table.setItem(r, c, item)

        self.table.setColumnWidth(0, 40)
        self.table.setColumnWidth(2, 110)
        self.table.setColumnWidth(3, 280)

        self.lbl_count.setText(f'기사 수: {len(results)}개')
        self.status.showMessage(
            f'✅  {len(results)}개 기사 수집 완료!  이제 [엑셀로 저장] 해보세요.')

    def _on_error(self, msg):
        self.pbar.setVisible(False)
        self.btn_crawl.setEnabled(True)
        self.lbl_count.setText('오류 발생')
        self.status.showMessage(f'❌  {msg}')
        QMessageBox.critical(self, '오류', f'{msg}\n\n네트워크를 확인하세요.')

    def _save_excel(self):
        if not self.results:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, '엑셀로 저장', '네이버_IT뉴스.xlsx', 'Excel 파일 (*.xlsx)')
        if not path:
            return
        try:
            save_to_excel(self.results, path)
            self.status.showMessage(f'💾  저장 완료: {path}')
            QMessageBox.information(self, '저장 완료',
                f'✅  {len(self.results)}개 기사 저장 완료!\n\n{path}')
        except Exception as e:
            QMessageBox.critical(self, '저장 오류', str(e))


# ── 실행 ─────────────────────────────────────────────────────

def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    pal = QPalette()
    pal.setColor(QPalette.ColorRole.Window,          QColor('#0d1117'))
    pal.setColor(QPalette.ColorRole.WindowText,      QColor('#e6edf3'))
    pal.setColor(QPalette.ColorRole.Base,            QColor('#161b22'))
    pal.setColor(QPalette.ColorRole.AlternateBase,   QColor('#1c2128'))
    pal.setColor(QPalette.ColorRole.Button,          QColor('#21262d'))
    pal.setColor(QPalette.ColorRole.ButtonText,      QColor('#e6edf3'))
    pal.setColor(QPalette.ColorRole.Highlight,       QColor('#1f6feb'))
    pal.setColor(QPalette.ColorRole.HighlightedText, QColor('#ffffff'))
    app.setPalette(pal)
    win = NewsApp()
    win.show()
    sys.exit(app.exec())

if __name__ == '__main__':
    main()
