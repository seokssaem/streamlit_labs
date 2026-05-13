"""
=============================================================
  📰 네이버 IT 뉴스 크롤러 + 엑셀 저장 + GUI 뷰어
  파이썬 기초반 종강 미니프로젝트
=============================================================
  ★ 지난 시간 코드와 비교해보세요! ★

  [지난 시간]                 [오늘]
  requests.get(url)       →  똑같이 사용
  BeautifulSoup(...)      →  똑같이 사용
  Workbook() / ws.append  →  똑같이 사용
  wb.save(파일명)          →  똑같이 사용
  print(결과)              →  GUI 테이블로 표시!

  설치:
    pip install PyQt6 requests beautifulsoup4 openpyxl
=============================================================
"""

import sys
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget,
    QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel,
    QTableWidget, QTableWidgetItem,
    QHeaderView, QStatusBar, QMessageBox,
    QFileDialog, QProgressBar,
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QColor, QPalette, QFont


# ══════════════════════════════════════════════════════════════
#  ★ 크롤링 함수 — 지난 시간 코드와 거의 동일! ★
# ══════════════════════════════════════════════════════════════

def crawl_naver_it_news():
    """
    지난 시간:
        url = 'https://news.naver.com/section/105'
        response = requests.get(url)
        soup = BeautifulSoup(response.text, 'html.parser')
        news = soup.find_all('strong', {'class': 'sa_text_strong'})
        for i, news_title in enumerate(news, start=1):
            result = {i: news_title.text}

    오늘: 딕셔너리 리스트로 모아서 반환!
    """
    url = 'https://news.naver.com/section/105'     # IT/과학 섹션 (지난 시간과 동일)

    response = requests.get(url, headers={
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'
    })
    html = response.text
    soup = BeautifulSoup(html, 'html.parser')

    # ── 지난 시간과 동일한 선택자 ──────────────────────────
    news_list = soup.find_all('strong', {'class': 'sa_text_strong'})

    results = []    # 지난 시간: result = {i: ...} → 오늘: 리스트에 딕셔너리를 담기

    for i, tag in enumerate(news_list, start=1):    # 지난 시간과 동일한 enumerate
        제목 = tag.text.strip()

        # 부모 <a> 태그에서 링크 꺼내기 (응용!)
        parent_a = tag.find_parent('a')
        링크 = parent_a['href'] if parent_a else ''

        # 카드 영역에서 언론사, 날짜 꺼내기 (응용!)
        card = tag.find_parent('li') or tag.find_parent('div')

        press_tag = card.find('a', {'class': 'sa_text_press'}) if card else None
        언론사 = press_tag.text.strip() if press_tag else ''

        date_tag = card.find('span', {'class': 'sa_text_datetime'}) if card else None
        날짜 = date_tag.text.strip() if date_tag else ''

        results.append({        # 지난 시간: {i: news_title.text} → 오늘: 컬럼명 붙이기!
            'N'  : i,
            '제목': 제목,
            '언론사': 언론사,
            '날짜': 날짜,
            '링크': 링크,
        })

    return results


# ══════════════════════════════════════════════════════════════
#  ★ 엑셀 저장 함수 — 지난 시간 코드와 거의 동일! ★
# ══════════════════════════════════════════════════════════════

def save_to_excel(results, filepath):
    """
    지난 시간 코드:
        wb = Workbook()
        ws = wb.active
        ws.title = '네이버 시가총액 상위'
        ws.append(['N', '종목명', '현재가', '시가총액'])
        for ... :
            ws.append([n, 종목명, 현재가, 시가총액])
        ws.column_dimensions['B'].width = 30
        wb.save('파일명.xlsx')
    """
    # ── 지난 시간과 똑같은 구조 ────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = '네이버 IT뉴스'                          # ← 시트 이름만 바꿈

    # 제목 행 추가 (지난 시간: ['N', '종목명', '현재가', '시가총액'])
    ws.append(['N', '제목', '언론사', '날짜', '링크'])

    # 제목 행 스타일 (응용!)
    header_fill = PatternFill('solid', fgColor='1F6FEB')
    header_font = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # 데이터 행 추가 (지난 시간: ws.append([n, 종목명, 현재가, 시가총액]))
    for row in results:
        ws.append([row['N'], row['제목'], row['언론사'], row['날짜'], row['링크']])

    # 열 너비 조정 (지난 시간과 동일!)
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 60    # 제목은 넓게
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 50

    # 엑셀 저장 (지난 시간과 동일!)
    wb.save(filepath)
    return len(results)


# ══════════════════════════════════════════════════════════════
#  백그라운드 워커 (크롤링 중 화면 안 멈추게)
# ══════════════════════════════════════════════════════════════

class CrawlWorker(QThread):
    finished = pyqtSignal(list)   # 크롤링 완료 → 결과 전달
    error    = pyqtSignal(str)    # 오류 발생 → 메시지 전달

    def run(self):
        try:
            results = crawl_naver_it_news()
            self.finished.emit(results)
        except Exception as e:
            self.error.emit(str(e))


# ══════════════════════════════════════════════════════════════
#  스타일
# ══════════════════════════════════════════════════════════════

STYLE = """
QMainWindow, QWidget {
    background-color: #0d1117;
    color: #e6edf3;
    font-family: "Malgun Gothic", "AppleGothic", sans-serif;
    font-size: 13px;
}
QPushButton {
    background: #21262d;
    color: #e6edf3;
    border: 1px solid #30363d;
    border-radius: 8px;
    padding: 10px 22px;
    font-size: 14px;
    font-weight: bold;
}
QPushButton:hover   { background:#30363d; border-color:#58a6ff; color:#58a6ff; }
QPushButton:pressed { background:#1f6feb; color:white; border-color:#1f6feb; }
QPushButton:disabled{ background:#161b22; color:#484f58; }

QPushButton#btn_crawl {
    background: #1f6feb;
    border: none; color: white;
    font-size: 15px;
    padding: 12px 36px;
    border-radius: 10px;
}
QPushButton#btn_crawl:hover    { background: #388bfd; }
QPushButton#btn_crawl:disabled { background: #0d419d; color: #8b949e; }

QPushButton#btn_save {
    background: #238636;
    border: none; color: white;
    border-radius: 8px; padding: 10px 24px;
}
QPushButton#btn_save:hover    { background: #2ea043; }
QPushButton#btn_save:disabled { background: #0f2d1a; color:#8b949e; }

QTableWidget {
    background: #161b22;
    border: 1px solid #30363d;
    border-radius: 8px;
    gridline-color: #21262d;
    color: #e6edf3;
}
QTableWidget::item:selected  { background: #1f6feb; color: white; }
QTableWidget::item:alternate { background: #1c2128; }
QHeaderView::section {
    background: #21262d;
    color: #58a6ff;
    padding: 8px;
    border: none;
    font-weight: bold;
}
QProgressBar {
    background: #21262d;
    border: none;
    border-radius: 4px;
    height: 6px;
}
QProgressBar::chunk { background: #1f6feb; border-radius: 4px; }
QStatusBar { background: #161b22; color: #8b949e; border-top: 1px solid #30363d; }
QScrollBar:vertical { background:#0d1117; width:8px; border-radius:4px; }
QScrollBar::handle:vertical { background:#30363d; border-radius:4px; min-height:24px; }
QScrollBar::handle:vertical:hover { background:#58a6ff; }
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height:0; }
"""


# ══════════════════════════════════════════════════════════════
#  메인 윈도우
# ══════════════════════════════════════════════════════════════

class NewsApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.results = []           # 크롤링 결과 보관

        self.setWindowTitle('📰 네이버 IT 뉴스 크롤러')
        self.setMinimumSize(1000, 650)
        self.setStyleSheet(STYLE)

        self._init_ui()

        self.status = QStatusBar()
        self.setStatusBar(self.status)
        self.status.showMessage('👋  [크롤링 시작] 버튼을 눌러 IT 뉴스를 가져오세요!')

    # ── UI 구성 ───────────────────────────────────────────────
    def _init_ui(self):
        center = QWidget()
        self.setCentralWidget(center)
        lay = QVBoxLayout(center)
        lay.setSpacing(12)
        lay.setContentsMargins(20, 16, 20, 16)

        # ① 헤더
        lay.addWidget(self._make_header())

        # ② 버튼 바
        lay.addWidget(self._make_btn_bar())

        # ③ 진행바 (크롤링 중에만 표시)
        self.pbar = QProgressBar()
        self.pbar.setVisible(False)
        self.pbar.setFixedHeight(6)
        self.pbar.setRange(0, 0)    # 무한 스피너
        lay.addWidget(self.pbar)

        # ④ 기사 수 표시
        self.lbl_count = QLabel('기사 수: —')
        self.lbl_count.setStyleSheet(
            'font-size:14px; font-weight:bold; color:#3fb950;')
        lay.addWidget(self.lbl_count)

        # ⑤ 결과 테이블
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(
            ['N', '제목', '언론사', '날짜', '링크'])
        self.table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch)      # 제목 열만 늘어나게
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(
            QTableWidget.SelectionBehavior.SelectRows)
        self.table.verticalHeader().setVisible(False)
        self.table.verticalHeader().setDefaultSectionSize(32)
        lay.addWidget(self.table, 1)

    # ── 헤더 ─────────────────────────────────────────────────
    def _make_header(self):
        w   = QWidget()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(0, 0, 0, 4)
        lay.setSpacing(2)

        t = QLabel('📰  네이버 IT 뉴스 크롤러')
        t.setStyleSheet('font-size:20px; font-weight:bold; color:#58a6ff;')

        s = QLabel(
            'requests + BeautifulSoup + openpyxl  →  PyQt6 GUI  ·  종강 미니프로젝트')
        s.setStyleSheet('font-size:12px; color:#8b949e;')

        lay.addWidget(t)
        lay.addWidget(s)
        return w

    # ── 버튼 바 ──────────────────────────────────────────────
    def _make_btn_bar(self):
        w   = QWidget()
        lay = QHBoxLayout(w)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(10)

        # 크롤링 버튼
        self.btn_crawl = QPushButton('🕷️  크롤링 시작')
        self.btn_crawl.setObjectName('btn_crawl')
        self.btn_crawl.clicked.connect(self._start_crawl)

        # 엑셀 저장 버튼
        self.btn_save = QPushButton('💾  엑셀로 저장')
        self.btn_save.setObjectName('btn_save')
        self.btn_save.setEnabled(False)
        self.btn_save.clicked.connect(self._save_excel)

        # 안내 라벨
        lbl = QLabel(
            '대상: 네이버 IT/과학 뉴스  (https://news.naver.com/section/105)')
        lbl.setStyleSheet('font-size:12px; color:#8b949e;')

        lay.addWidget(self.btn_crawl)
        lay.addWidget(self.btn_save)
        lay.addStretch()
        lay.addWidget(lbl)
        return w

    # ══════════════════════════════════════════════════════════
    #  크롤링
    # ══════════════════════════════════════════════════════════
    def _start_crawl(self):
        self.btn_crawl.setEnabled(False)
        self.btn_save.setEnabled(False)
        self.pbar.setVisible(True)
        self.table.setRowCount(0)
        self.lbl_count.setText('크롤링 중...')
        self.status.showMessage('🔄  네이버 IT 뉴스 가져오는 중...')

        self.worker = CrawlWorker()
        self.worker.finished.connect(self._on_done)
        self.worker.error.connect(self._on_error)
        self.worker.start()

    def _on_done(self, results):
        self.results = results
        self.pbar.setVisible(False)
        self.btn_crawl.setEnabled(True)
        self.btn_save.setEnabled(True)

        # ── 테이블에 결과 표시 ──────────────────────────────
        self.table.setRowCount(len(results))
        for r, row in enumerate(results):
            for c, key in enumerate(['N', '제목', '언론사', '날짜', '링크']):
                item = QTableWidgetItem(str(row[key]))
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                if key == 'N':
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.table.setItem(r, c, item)

        # 열 너비 (지난 시간 ws.column_dimensions 와 같은 개념)
        self.table.setColumnWidth(0, 40)   # N
        self.table.setColumnWidth(2, 110)  # 언론사
        self.table.setColumnWidth(3, 100)  # 날짜
        self.table.setColumnWidth(4, 280)  # 링크

        self.lbl_count.setText(f'기사 수: {len(results):,}개')
        self.status.showMessage(
            f'✅  {len(results)}개 기사 수집 완료!  [엑셀로 저장] 버튼으로 저장하세요.')

    def _on_error(self, msg):
        self.pbar.setVisible(False)
        self.btn_crawl.setEnabled(True)
        self.lbl_count.setText('오류 발생')
        self.status.showMessage(f'❌  오류: {msg}')
        QMessageBox.critical(self, '크롤링 오류',
            f'{msg}\n\n네트워크 연결을 확인하세요.')

    # ══════════════════════════════════════════════════════════
    #  엑셀 저장 — 지난 시간 wb.save() 와 동일한 흐름!
    # ══════════════════════════════════════════════════════════
    def _save_excel(self):
        if not self.results:
            return

        # 저장할 파일 경로를 대화상자로 선택
        path, _ = QFileDialog.getSaveFileName(
            self, '엑셀로 저장',
            '네이버_IT뉴스.xlsx',           # 기본 파일명
            'Excel 파일 (*.xlsx)'
        )
        if not path:
            return                          # 취소하면 아무것도 안 함

        try:
            count = save_to_excel(self.results, path)
            self.status.showMessage(f'💾  저장 완료: {path}')
            QMessageBox.information(self, '저장 완료',
                f'✅  {count}개 기사를 엑셀로 저장했습니다!\n\n{path}')
        except Exception as e:
            QMessageBox.critical(self, '저장 오류', str(e))


# ══════════════════════════════════════════════════════════════
#  실행
# ══════════════════════════════════════════════════════════════
def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')

    # 다크 팔레트
    pal = QPalette()
    pal.setColor(QPalette.ColorRole.Window,          QColor('#0d1117'))
    pal.setColor(QPalette.ColorRole.WindowText,      QColor('#e6edf3'))
    pal.setColor(QPalette.ColorRole.Base,            QColor('#161b22'))
    pal.setColor(QPalette.ColorRole.AlternateBase,   QColor('#1c2128'))
    pal.setColor(QPalette.ColorRole.Button,          QColor('#21262d'))
    pal.setColor(QPalette.ColorRole.ButtonText,      QColor('#e6edf3'))
    pal.setColor(QPalette.ColorRole.Highlight,       QColor('#1f6feb'))
    pal.setColor(QPalette.ColorRole.HighlightedText, QColor('#ffffff'))
    app.setPalette(pal)

    win = NewsApp()
    win.show()
    sys.exit(app.exec())

if __name__ == '__main__':
    main()
